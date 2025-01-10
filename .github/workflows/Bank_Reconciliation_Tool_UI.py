import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import sys
import subprocess

class BankReconApp:
    def __init__(self, root):
        self.root = root
        self.root.title("银行对账工具")
        
        # 设置窗口大小
        window_width = 610
        window_height = 500
        
        # 获取屏幕尺寸
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        
        # 计算居中位置
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # 设置窗口位置和大小
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # 创建主框架
        self.main_frame = ttk.Frame(root, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 文件选择部分
        self.file_frame = ttk.LabelFrame(self.main_frame, text="文件选择", padding="10")
        self.file_frame.pack(fill=tk.X, pady=5)
        
        self.gl_label = ttk.Label(self.file_frame, text="总账文件:")
        self.gl_label.grid(row=0, column=0, padx=5, pady=5)
        
        self.gl_entry = tk.Entry(self.file_frame, width=40)
        self.gl_entry.grid(row=0, column=1, padx=5, pady=5, sticky='ew')
        
        self.gl_button = tk.Button(self.file_frame, text="选择总账文件", command=self.select_gl_file)
        self.gl_button.grid(row=0, column=2, padx=5, pady=5)
        
        self.bank_label = ttk.Label(self.file_frame, text="银行文件:")
        self.bank_label.grid(row=1, column=0, padx=5, pady=5)
        
        self.bank_entry = tk.Entry(self.file_frame, width=40)
        self.bank_entry.grid(row=1, column=1, padx=5, pady=5, sticky='ew')
        
        self.bank_button = tk.Button(self.file_frame, text="选择银行文件", command=self.select_bank_file)
        self.bank_button.grid(row=1, column=2, padx=5, pady=5)
        
        # 操作按钮
        self.button_frame = ttk.Frame(self.main_frame)
        self.button_frame.pack(fill=tk.X, pady=10)
        
        self.process_button = ttk.Button(self.button_frame, text="开始对账", command=self.process_files)
        self.process_button.pack(side=tk.LEFT, padx=5)
        
        self.quit_button = ttk.Button(self.button_frame, text="退出", command=root.quit)
        self.quit_button.pack(side=tk.RIGHT, padx=5)
        
        # 日志显示
        self.log_frame = ttk.LabelFrame(self.main_frame, text="日志", padding="10")
        self.log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = tk.Text(self.log_frame, height=15)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(self.log_text)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.log_text.yview)
        
        # 添加底部标签
        self.bottom_label = tk.Label(root, text="Powered by Cayman FU", fg="gray")
        self.bottom_label.pack(side=tk.BOTTOM, pady=5)
        
    def select_gl_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[
                ("Excel files", "*.xlsx *.xls *.xlsm *.xlsb"),
                ("All files", "*.*")
            ]
        )
        if file_path:
            try:
                # 检查文件内容
                df = pd.read_excel(file_path, skiprows=1)
                required_columns = {'account', 'journal date', 'user', 'line description', 'base amount'}
                if not required_columns.issubset(df.columns.str.strip().str.lower()):
                    messagebox.showerror("错误", "总账文件缺少必要字段，请选择正确的文件")
                    return
                    
                self.gl_entry.delete(0, tk.END)
                self.gl_entry.insert(0, file_path)
                self.log(f"已选择总账文件: {file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"读取总账文件失败: {str(e)}")
    
    def select_bank_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[
                ("Excel files", "*.xls"),
                ("All files", "*.*")
            ]
        )
        if file_path:
            try:
                # 检查文件内容
                df = pd.read_excel(file_path, engine='xlrd', skiprows=8)
                required_columns = {'交易日期[ transaction date ]', '收款人名称[ payee\'s name ]', 
                                  '交易金额[ trade amount ]', '用途[ purpose ]'}
                if not required_columns.issubset(df.columns.str.strip().str.lower()):
                    messagebox.showerror("错误", "银行文件缺少必要字段，请选择正确的文件")
                    return
                    
                self.bank_entry.delete(0, tk.END)
                self.bank_entry.insert(0, file_path)
                self.log(f"已选择银行文件: {file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"读取银行文件失败: {str(e)}")
    
    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
    
    def clean_gl_data(self, file_path):
        df = pd.read_excel(file_path, sheet_name='sheet1', skiprows=1)
        df.columns = df.columns.str.strip().str.lower()
        
        if 'account' in df.columns:
            df['account'] = df['account'].astype(str).str.strip()
            filtered_df = df[df['account'] == '115307']
            
            new_columns = {
                'journal date': 'Date',
                'user': 'Reference',
                'line description': 'Description',
                'base amount': 'Base Amount'
            }
            cleaned_df = filtered_df[list(new_columns)].rename(columns=new_columns)
            cleaned_df['Date'] = pd.to_datetime(cleaned_df['Date']).dt.date
            
            return cleaned_df if not cleaned_df.empty else None
        return None

    def process_bank_data(self, file_path):
        df = pd.read_excel(file_path, engine='xlrd', skiprows=8)
        default_payee_name = "海南空港开发产业集团有限公司琼中福朋喜来登酒店分公司"
        
        def convert_date_format(date_str):
            try:
                return datetime.strptime(str(date_str), '%Y%m%d').strftime('%Y-%m-%d')
            except ValueError:
                return date_str
        
        new_rows = []
        for _, row in df.iterrows():
            transaction_date = convert_date_format(row.get('交易日期[ Transaction Date ]', ''))
            payee_name = row.get('收款人名称[ Payee\'s Name ]', np.nan)
            if pd.isna(payee_name) or str(payee_name).strip() == '':
                payee_name = default_payee_name
            
            trade_amount = float(row.get('交易金额[ Trade Amount ]', 0.0))
            debit_credit = "收款" if trade_amount > 0 else "付款" if trade_amount < 0 else ""
            
            new_row = {
                '日期': transaction_date,
                '对方户名': payee_name,
                '用途': row.get('用途[ Purpose ]', ''),
                '交易金额': trade_amount,
                '收款/付款': debit_credit,
                '交易流水号': str(row.get('交易流水号[ Transaction reference number ]', '')),
                '银行参考号': str(row.get('银行参考号', '')),
            }
            new_rows.append(new_row)
        
        return pd.DataFrame(new_rows)

    def match_data(self, gl_data, bank_data):
        gl_data['Base Amount'] = pd.to_numeric(gl_data['Base Amount'], errors='coerce')
        bank_data['交易金额'] = pd.to_numeric(bank_data['交易金额'], errors='coerce')

        matches = []
        unmatched_gl = []
        unmatched_bank = []

        bank_data_index_matched = set()
        gl_data_index_matched = set()

        for index_bank, row_bank in bank_data.iterrows():
            match = gl_data[
                (abs(gl_data['Base Amount']) == abs(row_bank['交易金额'])) & 
                (((row_bank['交易金额'] < 0) & (gl_data['Base Amount'] < 0)) | 
                 ((row_bank['交易金额'] > 0) & (gl_data['Base Amount'] > 0))) &
                (~gl_data.index.isin(gl_data_index_matched))
            ]
            
            if not match.empty:
                for index_gl, row_gl in match.iterrows():
                    formatted_date = row_gl['Date'].strftime('%Y-%m-%d') if not pd.isna(row_gl['Date']) else ''
                    
                    matches.append({
                        '日期': row_bank['日期'],
                        '对方户名': row_bank['对方户名'],
                        '用途': row_bank['用途'],
                        '交易流水号': str(row_bank['交易流水号']),
                        '收款/付款': row_bank['收款/付款'],
                        '交易金额': row_bank['交易金额'],
                        '与总帐核对': row_gl['Reference'],
                        ' ': '',  
                        'Check with Bank': str(row_bank['交易流水号']),
                        'Trans Date': formatted_date,
                        'Description': row_gl['Description'],
                        'Base Amount': row_gl['Base Amount'],
                    })
                    bank_data_index_matched.add(index_bank)
                    gl_data_index_matched.add(index_gl)
                    break  

        for index_gl, row_gl in gl_data.iterrows():
            if index_gl not in gl_data_index_matched:
                unmatched_gl.append({
                    'Trans Date': row_gl['Date'].strftime('%Y-%m-%d') if not pd.isna(row_gl['Date']) else '',
                    'Description': row_gl['Description'],
                    'Base Amount': row_gl['Base Amount'],
                    'Reference': row_gl['Reference']
                })

        for index_bank, row_bank in bank_data.iterrows():
            if index_bank not in bank_data_index_matched:
                unmatched_bank.append({
                    '日期': row_bank['日期'],
                    '对方户名': row_bank['对方户名'],
                    '用途': row_bank['用途'],
                    '收款/付款': row_bank['收款/付款'],
                    '交易金额': row_bank['交易金额'],
                    '交易流水号': str(row_bank['交易流水号'])
                })

        return matches, unmatched_gl, unmatched_bank

    def save_results(self, gl_data, bank_data, matches, unmatched_gl, unmatched_bank):
        output_file = 'Combined_Data.xlsx'
        
        # 创建新的Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Temp"
        wb.save(output_file)

        verify_data = pd.DataFrame(matches)
        unmatched_gl_df = pd.DataFrame(unmatched_gl)
        unmatched_bank_df = pd.DataFrame(unmatched_bank)

        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            verify_data.to_excel(writer, sheet_name='Bank_OK', index=False)
            unmatched_bank_df.to_excel(writer, sheet_name='Unmatched_Bank_Data', index=False)
            unmatched_gl_df.to_excel(writer, sheet_name='Unmatched_GL_Data', index=False)

        # 应用样式
        self.apply_excel_styles(output_file)
        
        # 获取完整路径
        full_path = os.path.abspath(output_file)
        
        # 更新日志信息
        self.log(f"处理完成！结果已保存到: {full_path}")
        
        # 询问用户是否要打开文件
        if messagebox.askyesno("处理完成", f"文件已保存到：\n{full_path}\n\n是否要立即打开？"):
            try:
                if os.name == 'nt':  # Windows
                    os.startfile(output_file)
                elif os.name == 'posix':  # macOS/Linux
                    if sys.platform == 'darwin':
                        subprocess.call(('open', output_file))
                    else:
                        subprocess.call(('xdg-open', output_file))
            except Exception as e:
                messagebox.showerror("错误", f"无法打开文件: {str(e)}")
        
        return output_file

    def apply_excel_styles(self, file_path):
        wb = load_workbook(file_path)

        # 定义样式
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100", bold=True)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        yellow_font = Font(color="000000", bold=True)
        header_fill = PatternFill(start_color="00009B", end_color="00009B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, name="微软雅黑", size=10)
        data_font = Font(name="微软雅黑", size=10)
        center_alignment = Alignment(horizontal='center')
        right_alignment = Alignment(horizontal='right')

        # 定义新的颜色样式
        special_header_fill = PatternFill(start_color="333F4F", end_color="333F4F", fill_type="solid")
        special_header_font = Font(color="FFFFFF", bold=True, name="微软雅黑", size=10)

        # 列宽和样式配置
        column_config = {
            'Bank_OK': {
                'widths': {
                    'A': 22.92, 'B': 42.26, 'C': 42.26, 'D': 17,
                    'E': 13.46, 'F': 13.46, 'G': 13.46, 'H': 2,
                    'I': 18, 'J': 22.92, 'K': 42.26, 'L': 13.46
                },
                'alignments': {
                    'A': center_alignment, 'B': center_alignment,
                    'C': center_alignment, 'D': center_alignment,
                    'E': center_alignment, 'F': right_alignment,
                    'G': center_alignment, 'H': center_alignment,
                    'I': center_alignment, 'J': center_alignment,
                    'K': center_alignment, 'L': right_alignment
                }
            },
            'Unmatched_GL_Data': {
                'widths': {
                    'A': 22.92, 'B': 42.26, 'C': 42.26, 'D': 13.46
                },
                'alignments': {
                    'A': center_alignment, 'B': center_alignment,
                    'C': right_alignment, 'D': right_alignment
                }
            },
            'Unmatched_Bank_Data': {
                'widths': {
                    'A': 22.92, 'B': 42.26, 'C': 42.26, 'D': 13.46,
                    'E': 13.46, 'F': 13.46
                },
                'alignments': {
                    'A': center_alignment, 'B': center_alignment,
                    'C': center_alignment, 'D': center_alignment,
                    'E': right_alignment, 'F': right_alignment
                }
            }
        }

        # 处理每个工作表
        for sheet_name in wb.sheetnames:
            if sheet_name in ['Temp', 'GL Data', 'Bank Data']:
                continue

            ws = wb[sheet_name]
            
            # 将生成器转换为列表以获取行数
            rows = list(ws.rows)
            
            # 如果工作表为空，跳过样式设置
            if len(rows) == 0:
                continue
            
            # 设置标题行
            if sheet_name == 'Bank_OK':
                ws.insert_rows(1)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(rows[0]))
                title_cell = ws['A1']
                title_cell.value = "银行数据 核对成功 明细"
                title_cell.font = green_font
                title_cell.fill = green_fill
                title_cell.alignment = center_alignment
            elif sheet_name == 'Unmatched_GL_Data':
                ws.insert_rows(1)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(rows[0]))
                title_cell = ws['A1']
                title_cell.value = "未匹配总帐数据"
                title_cell.font = yellow_font
                title_cell.fill = yellow_fill
                title_cell.alignment = center_alignment
            elif sheet_name == 'Unmatched_Bank_Data':
                ws.insert_rows(1)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(rows[0]))
                title_cell = ws['A1']
                title_cell.value = "未匹配银行数据"
                title_cell.font = yellow_font
                title_cell.fill = yellow_fill
                title_cell.alignment = center_alignment

            # 重新获取行数据，因为插入行后索引发生了变化
            rows = list(ws.rows)

            # 设置列宽和对齐方式
            if sheet_name in column_config:
                config = column_config[sheet_name]
                for col, width in config['widths'].items():
                    ws.column_dimensions[col].width = width
                    # 只对非空单元格应用样式
                    for cell in ws[col]:
                        if cell.value is not None:
                            cell.font = data_font
                            if col in config['alignments']:
                                cell.alignment = config['alignments'][col]

            # 设置表头样式（确保有足够行数）
            if len(rows) > 1:  # 现在第二行是表头行
                header_row = rows[1]  # 直接使用第二行
                for cell in header_row:
                    if cell.value is not None:  # 只对非空单元格应用样式
                        # 对I,J,K,L列应用特殊样式
                        if cell.column_letter in ['I', 'J', 'K', 'L'] and sheet_name == 'Bank_OK':
                            cell.fill = special_header_fill
                            cell.font = special_header_font
                        else:
                            cell.fill = header_fill
                            cell.font = header_font
                        cell.alignment = center_alignment

            # 冻结窗格（确保有足够行数）
            if len(rows) > 2:  # 从第三行开始冻结
                ws.freeze_panes = 'A3'

            # 设置会计专用格式（排除借方/贷方列）
            if sheet_name == 'Bank_OK':
                # F列（交易金额）和L列（Base Amount）设置会计格式
                for col in ['F', 'L']:
                    for cell in ws[col]:
                        if isinstance(cell.value, (int, float)):  # 只对数值单元格应用格式
                            cell.number_format = '#,##0.00_ ;[Red](#,##0.00)'
            elif sheet_name == 'Unmatched_Bank_Data':
                # E列（交易金额）设置会计格式
                for cell in ws['E']:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00_ ;[Red](#,##0.00)'
            elif sheet_name == 'Unmatched_GL_Data':
                # C列（Base Amount）设置会计格式
                for cell in ws['C']:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00_ ;[Red](#,##0.00)'

        # 隐藏不需要的工作表
        for sheet_name in ['Temp', 'GL Data', 'Bank Data']:
            if sheet_name in wb.sheetnames:
                wb[sheet_name].sheet_state = 'hidden'

        wb.save(file_path)

    def process_files(self):
        gl_file = self.gl_entry.get()
        bank_file = self.bank_entry.get()
        
        if not gl_file or not bank_file:
            messagebox.showerror("错误", "请先选择总账文件和银行文件")
            return
        
        try:
            self.log("开始处理文件...")
            
            # 清理总账数据
            self.log("清理总账数据...")
            gl_data = self.clean_gl_data(gl_file)
            if gl_data is None or gl_data.empty:
                raise ValueError("总账数据为空或格式不正确")
            
            # 处理银行数据
            self.log("处理银行数据...")
            bank_data = self.process_bank_data(bank_file)
            if bank_data.empty:
                raise ValueError("银行数据为空或格式不正确")
            
            # 进行数据匹配
            self.log("进行数据匹配...")
            matches, unmatched_gl, unmatched_bank = self.match_data(gl_data, bank_data)
            
            # 保存结果
            self.log("保存结果文件...")
            output_file = self.save_results(gl_data, bank_data, matches, unmatched_gl, unmatched_bank)
            

            
        except Exception as e:
            self.log(f"处理出错: {str(e)}")
            messagebox.showerror("错误", f"处理过程中出现错误: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = BankReconApp(root)
    root.mainloop() 
